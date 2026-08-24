from __future__ import annotations

import csv
import json
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List


def _in_range(row: Dict[str, Any], start: str, end: str) -> bool:
    value = str(row.get("date") or row.get("month") or "")[:10]
    return (not start or value >= start) and (not end or value <= end)


def _income_export_row(row: Dict[str, Any]) -> Dict[str, Any]:
    dine_in = float(row.get("dineIn", float(row.get("hall", 0) or 0) + float(row.get("room", 0) or 0)) or 0)
    chess = float(row.get("chess", 0) or 0)
    delivery = float(row.get("delivery", 0) or 0)
    date_value = str(row.get("date", ""))
    return {
        "记账日期": date_value,
        "录入方式": "按周期" if row.get("entryMode") == "period" else "按日",
        "周期开始": row.get("periodStart") or date_value,
        "周期结束": row.get("periodEnd") or date_value,
        "堂食": dine_in,
        "棋牌房": chess,
        "外送": delivery,
        "合计": dine_in + chess + delivery,
        "备注": row.get("note", ""),
    }


def datasets(state: Dict[str, Any], start: str = "", end: str = "") -> Dict[str, List[Dict[str, Any]]]:
    names = ["incomeRecords", "salesRecords", "expenses", "products", "stocktakes", "reminders", "employees", "payrolls", "suppliers", "assets"]
    result: Dict[str, List[Dict[str, Any]]] = {}
    for name in names:
        rows = list(state.get(name, []))
        if start or end:
            rows = [row for row in rows if not ("date" in row or "month" in row) or _in_range(row, start, end)]
        if name == "incomeRecords":
            rows = [_income_export_row(row) for row in rows]
        result[name] = rows
    return result


def export_csv_zip(state: Dict[str, Any], target: Path, start: str = "", end: str = "") -> Path:
    with tempfile.TemporaryDirectory() as temp:
        temp_dir = Path(temp)
        files = []
        for name, rows in datasets(state, start, end).items():
            path = temp_dir / f"{name}.csv"
            fields = sorted({key for row in rows for key in row.keys()}) or ["无数据"]
            with path.open("w", newline="", encoding="utf-8-sig") as handle:
                writer = csv.DictWriter(handle, fieldnames=fields)
                writer.writeheader()
                for row in rows:
                    writer.writerow({key: json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value for key, value in row.items()})
            files.append(path)
        with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
            for path in files:
                archive.write(path, path.name)
    return target


def export_xlsx(state: Dict[str, Any], target: Path, start: str = "", end: str = "") -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    book = Workbook()
    book.remove(book.active)
    for name, rows in datasets(state, start, end).items():
        sheet = book.create_sheet(name[:31])
        fields = sorted({key for row in rows for key in row.keys()}) or ["无数据"]
        sheet.append(fields)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="147D68")
        for row in rows:
            sheet.append([json.dumps(row.get(key), ensure_ascii=False) if isinstance(row.get(key), (dict, list)) else row.get(key) for key in fields])
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(45, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    book.save(target)
    return target
