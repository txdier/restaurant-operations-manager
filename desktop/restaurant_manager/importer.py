from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


QUICK_HEADERS = ["日期", "支出类别", "金额", "经手人", "备注"]
DETAIL_HEADERS = ["采购单号", "日期", "商品名称", "支出类别", "数量", "单位", "单价", "经手人", "备注"]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _day(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = _text(value)
    if not text:
        return ""
    return datetime.strptime(text[:10].replace("/", "-"), "%Y-%m-%d").strftime("%Y-%m-%d")


def _number(value: Any, field: str) -> float:
    try:
        number = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field}不是有效数字") from None
    if number <= 0:
        raise ValueError(f"{field}必须大于 0")
    return float(number)


def _money_key(value: Any) -> int:
    number = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return int(number * 100)


def _quick_key(row: Dict[str, Any]) -> tuple[str, str, int, str, str]:
    return (
        _text(row.get("date")),
        _text(row.get("category")),
        _money_key(row.get("amount", 0)),
        _text(row.get("handler")),
        _text(row.get("note")),
    )


def _expense_quick_key(expense: Dict[str, Any]) -> tuple[str, str, int, str, str] | None:
    if expense.get("purchaseNo") or _text(expense.get("mode")) == "详细采购":
        return None
    category = _text(expense.get("category"))
    item = _text(expense.get("item"))
    note = _text(expense.get("note"))
    if not note and item not in (category, "快速支出"):
        note = item
    try:
        return (
            _text(expense.get("date")),
            category,
            _money_key(expense.get("amount", 0)),
            _text(expense.get("handler")),
            note,
        )
    except (InvalidOperation, ValueError, TypeError):
        return None


def create_import_template(target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    quick = book.active
    quick.title = "快速支出"
    detail = book.create_sheet("详细采购")
    notes = book.create_sheet("填写说明")
    for sheet, headers in ((quick, QUICK_HEADERS), (detail, DETAIL_HEADERS)):
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F6FED")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 16
    quick.append([date.today(), "水电燃气", 680, "李经理", "本月水电费"])
    detail.append(["CG-202608-001", date.today(), "五花肉", "食材", 12, "kg", 28, "张师傅", "晨间采购"])
    detail.append(["CG-202608-001", date.today(), "土豆", "食材", 20, "kg", 3.2, "张师傅", "同一采购单"])
    notes.append(["工作表", "用途", "关键规则"])
    notes.append(["快速支出", "不需要商品、数量、单价的费用", "日期、类别、金额、经手人、备注完全相同时视为重复，默认跳过"])
    notes.append(["详细采购", "按商品逐行记录采购明细", "相同采购单号会合并为一张采购单；商品名称+单位用于匹配商品"])
    notes.append(["通用", "日期格式 YYYY-MM-DD", "支出类别必须已在系统中启用；不要修改工作表名称和表头"])
    book.save(target)
    return target


def preview_import(path: Path, state: Dict[str, Any]) -> Dict[str, Any]:
    if path.suffix.lower() != ".xlsx":
        raise ValueError("导入文件必须是 .xlsx 格式")
    if not path.is_file():
        raise ValueError(f"找不到所选文件：{path}。请重新选择；Windows 7 建议先将文件放到桌面后再试")
    book = load_workbook(path, data_only=True, read_only=True)
    missing = [name for name in ("快速支出", "详细采购") if name not in book.sheetnames]
    if missing:
        raise ValueError("缺少工作表：" + "、".join(missing))
    active_categories = {c["name"] for c in state.get("expenseCategories", []) if c.get("active", True)}
    products = state.get("products", [])
    product_pairs = {(_text(p.get("name")), _text(p.get("unit"))): p for p in products if p.get("active", True)}
    units_by_name: Dict[str, set[str]] = {}
    for product in products:
        units_by_name.setdefault(_text(product.get("name")), set()).add(_text(product.get("unit")))
    existing_orders = {_text(e.get("purchaseNo")) for e in state.get("expenses", []) if e.get("purchaseNo")}
    existing_quick = {key for expense in state.get("expenses", []) if (key := _expense_quick_key(expense)) is not None}
    errors: list[str] = []
    warnings: list[str] = []
    quick_rows: list[Dict[str, Any]] = []
    duplicate_quick_rows: list[Dict[str, Any]] = []
    seen_quick: set[tuple[str, str, int, str, str]] = set()
    purchase_map: Dict[str, Dict[str, Any]] = {}
    unknown: Dict[tuple[str, str], Dict[str, str]] = {}

    def rows(sheet_name: str, headers: list[str]):
        sheet = book[sheet_name]
        actual = [_text(v) for v in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        if actual[: len(headers)] != headers:
            errors.append(f"{sheet_name}：表头不正确，请使用系统模板")
            return []
        return list(sheet.iter_rows(min_row=2, values_only=True))

    for index, values in enumerate(rows("快速支出", QUICK_HEADERS), 2):
        if not any(v not in (None, "") for v in values):
            continue
        try:
            day, category, handler, note = _day(values[0]), _text(values[1]), _text(values[3]), _text(values[4])
            if not day or not category:
                raise ValueError("日期和支出类别不能为空")
            if category not in active_categories:
                raise ValueError(f"支出类别“{category}”不存在或已停用")
            row = {"date": day, "category": category, "amount": _number(values[2], "金额"), "handler": handler, "note": note}
            key = _quick_key(row)
            if key in existing_quick or key in seen_quick:
                duplicate_quick_rows.append({**row, "row": index, "reason": "系统已有相同记录" if key in existing_quick else "文件内重复"})
            else:
                quick_rows.append(row)
                seen_quick.add(key)
        except (ValueError, TypeError) as error:
            errors.append(f"快速支出第 {index} 行：{error}")

    for index, values in enumerate(rows("详细采购", DETAIL_HEADERS), 2):
        if not any(v not in (None, "") for v in values):
            continue
        try:
            purchase_no, day, name, category = _text(values[0]), _day(values[1]), _text(values[2]), _text(values[3])
            unit, handler, note = _text(values[5]), _text(values[7]), _text(values[8])
            if not purchase_no or not day or not name or not category or not unit:
                raise ValueError("采购单号、日期、商品名称、支出类别和单位不能为空")
            if category not in active_categories:
                raise ValueError(f"支出类别“{category}”不存在或已停用")
            if purchase_no in existing_orders:
                raise ValueError(f"采购单号“{purchase_no}”已导入，不能重复导入")
            product = product_pairs.get((name, unit))
            if not product and name in units_by_name and unit not in units_by_name[name]:
                raise ValueError(f"商品“{name}”已存在，但单位不同；请先确认商品单位")
            if not product:
                unknown[(name, unit)] = {"name": name, "unit": unit, "category": category}
            line = {"productId": product.get("id") if product else None, "name": name, "category": category, "qty": _number(values[4], "数量"), "unit": unit, "price": _number(values[6], "单价"), "note": note}
            purchase = purchase_map.setdefault(purchase_no, {"purchaseNo": purchase_no, "date": day, "handler": handler, "lines": []})
            if purchase["date"] != day or purchase["handler"] != handler:
                raise ValueError(f"同一采购单号的日期和经手人必须一致（{purchase_no}）")
            purchase["lines"].append(line)
        except (ValueError, TypeError) as error:
            errors.append(f"详细采购第 {index} 行：{error}")
    purchases = list(purchase_map.values())
    if unknown:
        warnings.append(f"有 {len(unknown)} 个商品尚未建立，可在导入时自动新增")
    if duplicate_quick_rows:
        warnings.append(f"发现 {len(duplicate_quick_rows)} 笔重复快速支出，默认跳过；如确认是不同业务，可选择仍然导入")
    return {"path": str(path), "quickExpenses": quick_rows, "duplicateQuickExpenses": duplicate_quick_rows, "purchases": purchases, "unknownProducts": list(unknown.values()), "errors": errors, "warnings": warnings, "counts": {"quickExpenses": len(quick_rows), "duplicateQuickExpenses": len(duplicate_quick_rows), "purchaseOrders": len(purchases), "purchaseLines": sum(len(p["lines"]) for p in purchases)}}


def apply_import(preview: Dict[str, Any], state: Dict[str, Any], create_unknown_products: bool, import_duplicate_quick_expenses: bool = False) -> Dict[str, Any]:
    if preview.get("errors"):
        raise ValueError("导入预览仍有错误，请先修正 Excel 文件")
    if preview.get("unknownProducts") and not create_unknown_products:
        raise ValueError("存在未建立商品，请勾选自动新增商品或先在商品管理中建立")
    batch_id = uuid4().hex
    next_product_id = max([int(p.get("id", 0)) for p in state.get("products", [])] + [0]) + 1
    product_map = {(_text(p.get("name")), _text(p.get("unit"))): p for p in state.get("products", [])}
    for unknown in preview.get("unknownProducts", []):
        product = {"id": next_product_id, "name": unknown["name"], "category": unknown["category"], "brand": "", "spec": "", "unit": unknown["unit"], "stocktake": False, "reminder": False, "active": True, "createdAt": date.today().isoformat()}
        state.setdefault("products", []).append(product)
        product_map[(product["name"], product["unit"])] = product
        next_product_id += 1
    next_expense_id = max([int(e.get("id", 0)) for e in state.get("expenses", [])] + [0]) + 1
    added: list[Dict[str, Any]] = []
    quick_rows = list(preview.get("quickExpenses", []))
    if import_duplicate_quick_expenses:
        quick_rows.extend(preview.get("duplicateQuickExpenses", []))
    for row in quick_rows:
        added.append({"id": next_expense_id, "date": row["date"], "mode": "快速记账", "category": row["category"], "item": row.get("note") or row["category"], "amount": row["amount"], "handler": row["handler"], "status": "有效", "importBatchId": batch_id})
        next_expense_id += 1
    for purchase in preview.get("purchases", []):
        for line in purchase["lines"]:
            product = product_map[(line["name"], line["unit"])]
            amount = float((Decimal(str(line["qty"])) * Decimal(str(line["price"]))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            added.append({"id": next_expense_id, "date": purchase["date"], "mode": "详细采购", "category": line["category"], "item": f'{line["name"]} {line["qty"]:g}{line["unit"]}', "amount": amount, "handler": purchase["handler"], "status": "有效", "purchaseNo": purchase["purchaseNo"], "productId": product["id"], "qty": line["qty"], "unit": line["unit"], "price": line["price"], "note": line.get("note", ""), "importBatchId": batch_id})
            next_expense_id += 1
    state.setdefault("expenses", []).extend(added)
    state.setdefault("importBatches", []).append({"id": batch_id, "file": Path(preview["path"]).name, "importedAt": datetime.now().isoformat(timespec="seconds"), "quickExpenses": len(quick_rows), "duplicateQuickExpenses": len(preview.get("duplicateQuickExpenses", [])) if import_duplicate_quick_expenses else 0, "purchaseOrders": preview["counts"]["purchaseOrders"], "purchaseLines": preview["counts"]["purchaseLines"]})
    return state
