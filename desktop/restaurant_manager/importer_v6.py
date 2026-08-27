from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple
from uuid import uuid4

from .money import cents_to_legacy_number, yuan_to_cents
from .storage_v6 import _uid


QUICK_HEADERS = ["日期", "支出类别", "金额", "经手人", "备注"]
DETAIL_HEADERS = ["采购单号", "日期", "商品名称", "支出类别", "数量", "单位", "单价", "经手人", "备注"]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _day(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = _text(value)
    if not text:
        return ""
    return datetime.strptime(text[:10].replace("/", "-"), "%Y-%m-%d").strftime("%Y-%m-%d")


def _quantity(value: Any, field: str) -> Decimal:
    try:
        number = Decimal(str(value)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field}不是有效数字") from None
    if number <= 0:
        raise ValueError(f"{field}必须大于 0")
    return number


def _display_decimal(value: Decimal) -> float | int:
    return int(value) if value == value.to_integral_value() else float(value)


def _quick_key(row: Dict[str, Any]) -> Tuple[str, str, int, str, str]:
    return (
        _text(row.get("date")),
        _text(row.get("category")),
        int(row.get("amountCents", 0)),
        _text(row.get("handler")),
        _text(row.get("note")),
    )


def _quick_identity_key(row: Dict[str, Any]) -> Tuple[str, int, str, str]:
    return (
        _text(row.get("date")),
        int(row.get("amountCents", 0)),
        _text(row.get("handler")),
        _text(row.get("note")) or _text(row.get("category")),
    )


def _system_quick_keys(conn, row: Dict[str, Any]) -> tuple[set[Tuple[str, str, int, str, str]], set[Tuple[str, int, str, str]]]:
    records = conn.execute(
        """SELECT category_name_snapshot,item,note
           FROM expenses_v6
           WHERE expense_date=? AND amount_cents=? AND handler=?
             AND COALESCE(purchase_no,'')='' AND mode<>'详细采购'""",
        (row["date"], row["amountCents"], row["handler"]),
    ).fetchall()
    full: set[Tuple[str, str, int, str, str]] = set()
    identities: set[Tuple[str, int, str, str]] = set()
    for category, item, note in records:
        category_text = _text(category)
        item_text = _text(item)
        note_text = _text(note)
        if not note_text and item_text not in (category_text, "快速支出"):
            note_text = item_text
        full.add((row["date"], category_text, row["amountCents"], row["handler"], note_text))
        identities.add((row["date"], row["amountCents"], row["handler"], note_text or category_text))
    return full, identities


def preview_import_v6(path: Path, database: Any) -> Dict[str, Any]:
    if path.suffix.lower() != ".xlsx":
        raise ValueError("导入文件必须是 .xlsx 格式")
    if not path.is_file():
        raise ValueError(f"找不到所选文件：{path}。请重新选择；Windows 7 建议先将文件放到桌面后再试")

    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True, read_only=True)
    try:
        missing = [name for name in ("快速支出", "详细采购") if name not in book.sheetnames]
        if missing:
            raise ValueError("缺少工作表：" + "、".join(missing))

        with database.lock, database.connect() as conn:
            active_categories = {str(row[0]) for row in conn.execute("SELECT name FROM expense_categories_v6 WHERE active=1")}
            product_rows = conn.execute("SELECT id,name,unit,category_name_snapshot,active FROM products_v6 ORDER BY id").fetchall()
            product_pairs = {(str(row[1]), str(row[2])): {"id": int(row[0]), "name": str(row[1]), "unit": str(row[2]), "category": str(row[3])} for row in product_rows if row[4]}
            inactive_product_pairs = {(str(row[1]), str(row[2])) for row in product_rows if not row[4]}
            units_by_name: Dict[str, set[str]] = {}
            for row in product_rows:
                units_by_name.setdefault(str(row[1]), set()).add(str(row[2]))

            errors: List[str] = []
            warnings: List[str] = []
            quick_rows: List[Dict[str, Any]] = []
            duplicate_quick_rows: List[Dict[str, Any]] = []
            seen_quick: set[Tuple[str, str, int, str, str]] = set()
            purchase_map: Dict[str, Dict[str, Any]] = {}
            unknown: Dict[Tuple[str, str], Dict[str, str]] = {}

            def rows(sheet_name: str, headers: List[str]) -> Iterable[tuple[Any, ...]]:
                sheet = book[sheet_name]
                actual = [_text(v) for v in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                if actual[: len(headers)] != headers:
                    errors.append(f"{sheet_name}：表头不正确，请使用系统模板")
                    return []
                return sheet.iter_rows(min_row=2, values_only=True)

            for index, values in enumerate(rows("快速支出", QUICK_HEADERS), 2):
                if not any(v not in (None, "") for v in values):
                    continue
                try:
                    day, category, handler, note = _day(values[0]), _text(values[1]), _text(values[3]), _text(values[4])
                    if not day or not category:
                        raise ValueError("日期和支出类别不能为空")
                    if category not in active_categories:
                        raise ValueError(f"支出类别“{category}”不存在或已停用")
                    amount_cents = yuan_to_cents(values[2])
                    if amount_cents <= 0:
                        raise ValueError("金额必须大于 0")
                    row = {
                        "date": day,
                        "category": category,
                        "amount": cents_to_legacy_number(amount_cents),
                        "amountCents": amount_cents,
                        "handler": handler,
                        "note": note,
                    }
                    system_full, system_identity = _system_quick_keys(conn, row)
                    key = _quick_key(row)
                    exists_in_system = key in system_full or _quick_identity_key(row) in system_identity
                    if exists_in_system or key in seen_quick:
                        duplicate_quick_rows.append({**row, "row": index, "reason": "系统已有相同记录" if exists_in_system else "文件内重复"})
                    else:
                        quick_rows.append(row)
                        seen_quick.add(key)
                except (ValueError, TypeError) as error:
                    errors.append(f"快速支出第 {index} 行：{error}")

            for index, values in enumerate(rows("详细采购", DETAIL_HEADERS), 2):
                if not any(v not in (None, "") for v in values):
                    continue
                try:
                    purchase_no, day, name, category = _text(values[0]), _day(values[1]), _text(values[2]), _text(values[3])
                    unit, handler, note = _text(values[5]), _text(values[7]), _text(values[8])
                    if not purchase_no or not day or not name or not category or not unit:
                        raise ValueError("采购单号、日期、商品名称、支出类别和单位不能为空")
                    if category not in active_categories:
                        raise ValueError(f"支出类别“{category}”不存在或已停用")
                    if conn.execute("SELECT 1 FROM expenses_v6 WHERE purchase_no=? LIMIT 1", (purchase_no,)).fetchone():
                        raise ValueError(f"采购单号“{purchase_no}”已导入，不能重复导入")
                    product = product_pairs.get((name, unit))
                    if not product and (name, unit) in inactive_product_pairs:
                        raise ValueError(f"商品“{name}（{unit}）”已停用，请先在商品管理中启用后再导入")
                    if not product and name in units_by_name and unit not in units_by_name[name]:
                        raise ValueError(f"商品“{name}”已存在，但单位不同；请先确认商品单位")
                    if not product:
                        unknown[(name, unit)] = {"name": name, "unit": unit, "category": category}
                    qty_decimal = _quantity(values[4], "数量")
                    price_cents = yuan_to_cents(values[6])
                    if price_cents <= 0:
                        raise ValueError("单价必须大于 0")
                    line = {
                        "productId": product.get("id") if product else None,
                        "name": name,
                        "category": category,
                        "qty": _display_decimal(qty_decimal),
                        "qtyDecimal": str(qty_decimal),
                        "unit": unit,
                        "price": cents_to_legacy_number(price_cents),
                        "priceCents": price_cents,
                        "note": note,
                    }
                    purchase = purchase_map.setdefault(purchase_no, {"purchaseNo": purchase_no, "date": day, "handler": handler, "lines": []})
                    if purchase["date"] != day or purchase["handler"] != handler:
                        raise ValueError(f"同一采购单号的日期和经手人必须一致（{purchase_no}）")
                    purchase["lines"].append(line)
                except (ValueError, TypeError) as error:
                    errors.append(f"详细采购第 {index} 行：{error}")

        purchases = list(purchase_map.values())
        if unknown:
            warnings.append(f"有 {len(unknown)} 个商品尚未建立，可在导入时自动新增")
        if duplicate_quick_rows:
            warnings.append(f"发现 {len(duplicate_quick_rows)} 笔重复快速支出，默认跳过；如确认是不同业务，可选择仍然导入")
        return {
            "path": str(path),
            "quickExpenses": [{k: v for k, v in row.items() if k != "amountCents"} for row in quick_rows],
            "duplicateQuickExpenses": [{k: v for k, v in row.items() if k != "amountCents"} for row in duplicate_quick_rows],
            "purchases": [{**purchase, "lines": [{k: v for k, v in line.items() if k not in ("priceCents", "qtyDecimal")} for line in purchase["lines"]]} for purchase in purchases],
            "unknownProducts": list(unknown.values()),
            "errors": errors,
            "warnings": warnings,
            "counts": {
                "quickExpenses": len(quick_rows),
                "duplicateQuickExpenses": len(duplicate_quick_rows),
                "purchaseOrders": len(purchases),
                "purchaseLines": sum(len(p["lines"]) for p in purchases),
            },
            "_internal": {
                "quickExpenses": quick_rows,
                "duplicateQuickExpenses": duplicate_quick_rows,
                "purchases": purchases,
            },
        }
    finally:
        book.close()


def _category_ids(conn) -> Dict[str, int]:
    return {str(name): int(row_id) for row_id, name in conn.execute("SELECT id,name FROM expense_categories_v6 WHERE active=1")}


def _product_pairs(conn) -> Dict[Tuple[str, str], Dict[str, Any]]:
    return {
        (str(name), str(unit)): {"id": int(row_id), "name": str(name), "unit": str(unit), "category": str(category)}
        for row_id, name, unit, category in conn.execute("SELECT id,name,unit,category_name_snapshot FROM products_v6 WHERE active=1")
    }


def apply_import_v6(
    preview: Dict[str, Any],
    database: Any,
    create_unknown_products: bool,
    import_duplicate_quick_expenses: bool = False,
) -> Dict[str, Any]:
    if preview.get("errors"):
        raise ValueError("导入预览仍有错误，请先修正 Excel 文件")
    if preview.get("unknownProducts") and not create_unknown_products:
        raise ValueError("存在未建立商品，请勾选自动新增商品或先在商品管理中建立")

    internal = preview.get("_internal") or {}
    quick_rows = list(internal.get("quickExpenses", []))
    if import_duplicate_quick_expenses:
        quick_rows.extend(internal.get("duplicateQuickExpenses", []))
    purchases = list(internal.get("purchases", []))
    batch_id = uuid4().hex
    imported_at = datetime.now().isoformat(timespec="seconds")

    with database.lock, database.connect() as conn:
        conn.execute("BEGIN IMMEDIATE")
        category_ids = _category_ids(conn)
        product_pairs = _product_pairs(conn)
        required_categories = {
            _text(row.get("category")) for row in quick_rows
        } | {
            _text(line.get("category")) for purchase in purchases for line in purchase.get("lines", [])
        }
        unavailable_categories = sorted(name for name in required_categories if name not in category_ids)
        if unavailable_categories:
            raise ValueError("以下支出类别不存在或已停用：" + "、".join(unavailable_categories))
        duplicate_orders = sorted({
            str(purchase.get("purchaseNo", ""))
            for purchase in purchases
            if conn.execute(
                "SELECT 1 FROM expenses_v6 WHERE purchase_no=? LIMIT 1",
                (str(purchase.get("purchaseNo", "")),),
            ).fetchone()
        })
        if duplicate_orders:
            raise ValueError("以下采购单号已存在，不能重复导入：" + "、".join(duplicate_orders))
        next_product_id = int(conn.execute("SELECT COALESCE(MAX(id),0)+1 FROM products_v6").fetchone()[0])
        for unknown in preview.get("unknownProducts", []):
            pair = (_text(unknown.get("name")), _text(unknown.get("unit")))
            if pair in product_pairs:
                continue
            product = {"id": next_product_id, "name": pair[0], "unit": pair[1], "category": _text(unknown.get("category"))}
            conn.execute(
                """INSERT INTO products_v6(id,uid,name,category_name_snapshot,brand,spec,unit,stocktake_enabled,reminder_enabled,active,created_at,legacy_json)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (next_product_id, _uid("product", next_product_id, str(next_product_id)), pair[0], product["category"], "", "", pair[1], 0, 0, 1, date.today().isoformat(), "{}"),
            )
            product_pairs[pair] = product
            next_product_id += 1

        next_expense_id = int(conn.execute("SELECT COALESCE(MAX(id),0)+1 FROM expenses_v6").fetchone()[0])
        added_ids: List[int] = []
        for row in quick_rows:
            row_id = next_expense_id
            next_expense_id += 1
            conn.execute(
                """INSERT INTO expenses_v6(id,uid,expense_date,mode,category_id,category_name_snapshot,item,amount_cents,handler,status,note,import_batch_id,legacy_json)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    row_id, _uid("expense", row_id, str(row_id)), row["date"], "快速记账", category_ids.get(row["category"]), row["category"],
                    row.get("note") or row["category"], int(row["amountCents"]), row["handler"], "有效", row.get("note", ""), batch_id, "{}",
                ),
            )
            added_ids.append(row_id)

        purchase_orders: set[str] = set()
        for purchase in purchases:
            purchase_orders.add(str(purchase["purchaseNo"]))
            for line in purchase["lines"]:
                pair = (_text(line["name"]), _text(line["unit"]))
                product = product_pairs.get(pair)
                if not product:
                    raise ValueError(f"商品“{pair[0]}（{pair[1]}）”未建立或已停用")
                qty = Decimal(str(line["qtyDecimal"]))
                price_cents = int(line["priceCents"])
                total_yuan = (qty * (Decimal(price_cents) / Decimal(100))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                total_cents = int(total_yuan * 100)
                row_id = next_expense_id
                next_expense_id += 1
                conn.execute(
                    """INSERT INTO expenses_v6(id,uid,expense_date,mode,category_id,category_name_snapshot,item,amount_cents,handler,status,note,purchase_no,product_id,product_name_snapshot,quantity,unit_snapshot,unit_price_cents,import_batch_id,legacy_json)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        row_id, _uid("expense", row_id, str(row_id)), purchase["date"], "详细采购", category_ids.get(line["category"]), line["category"],
                        f'{line["name"]} {_display_decimal(qty):g}{line["unit"]}', total_cents, purchase["handler"], "有效", line.get("note", ""),
                        purchase["purchaseNo"], product["id"], product["name"], float(qty), line["unit"], price_cents, batch_id, "{}",
                    ),
                )
                added_ids.append(row_id)

        batch_payload = {
            "id": batch_id,
            "file": Path(preview["path"]).name,
            "importedAt": imported_at,
            "quickExpenses": len(quick_rows),
            "duplicateQuickExpenses": len(internal.get("duplicateQuickExpenses", [])) if import_duplicate_quick_expenses else 0,
            "purchaseOrders": len(purchase_orders),
            "purchaseLines": sum(len(p.get("lines", [])) for p in purchases),
        }
        conn.execute(
            "INSERT INTO import_batches_v6(id,file_name,imported_at,payload_json) VALUES(?,?,?,?)",
            (batch_id, batch_payload["file"], imported_at, json.dumps(batch_payload, ensure_ascii=False, separators=(",", ":"))),
        )

        conn.execute(
            "INSERT INTO audit_log(event,detail) VALUES(?,?)",
            ("import_expenses_v6", json.dumps({"batchId": batch_id, "expenseIds": added_ids}, ensure_ascii=False, separators=(",", ":"))),
        )
        conn.commit()

    return {"batchId": batch_id, "counts": batch_payload}


def public_preview(preview: Dict[str, Any]) -> Dict[str, Any]:
    return {key: value for key, value in preview.items() if key != "_internal"}
