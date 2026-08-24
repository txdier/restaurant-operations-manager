import zipfile
from pathlib import Path

from openpyxl import load_workbook

from restaurant_manager.exporter import datasets, export_csv_zip, export_query_csv, export_xlsx
from restaurant_manager.migrations import default_state


def test_exports_all_business_sections(tmp_path: Path):
    state = default_state()
    state["expenses"] = [{"id": 1, "date": "2026-08-19", "amount": 12.5, "item": "测试"}]
    xlsx = export_xlsx(state, tmp_path / "all.xlsx")
    assert "expenses" in load_workbook(xlsx, read_only=True).sheetnames
    archive = export_csv_zip(state, tmp_path / "all.zip")
    with zipfile.ZipFile(archive) as zf:
        assert "expenses.csv" in zf.namelist()


def test_income_export_merges_historical_hall_and_room_without_rewriting_source():
    state = default_state()
    historical = {"id": 1, "date": "2026-08-19", "hall": 3200, "room": 2100, "chess": 600, "delivery": 750, "note": "历史"}
    state["incomeRecords"] = [historical]
    exported = datasets(state)["incomeRecords"][0]
    assert exported["堂食"] == 5300
    assert exported["合计"] == 6650
    assert "hall" not in exported and "room" not in exported
    assert historical["hall"] == 3200 and historical["room"] == 2100


def test_period_income_export_uses_period_end_as_accounting_date():
    state = default_state()
    state["incomeRecords"] = [{"id": 2, "date": "2026-09-01", "entryMode": "period", "periodStart": "2026-08-26", "periodEnd": "2026-09-01", "dineIn": 20000, "chess": 0, "delivery": 0}]
    assert datasets(state, "2026-09-01", "2026-09-30")["incomeRecords"][0]["录入方式"] == "按周期"
    assert datasets(state, "2026-08-01", "2026-08-31")["incomeRecords"] == []


def test_query_csv_export_writes_utf8_bom_and_quoted_content(tmp_path: Path):
    target = export_query_csv(
        ["日期", "项目", "金额"],
        [["2026-08-24", '米油,含"税"', 128.5]],
        tmp_path / "query.csv",
    )

    raw = target.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")
    text = raw.decode("utf-8-sig")
    assert "日期,项目,金额" in text
    assert '"米油,含""税"""' in text
