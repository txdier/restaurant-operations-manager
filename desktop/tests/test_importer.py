from pathlib import Path

from openpyxl import load_workbook

from restaurant_manager.importer import apply_import, create_import_template, preview_import
from restaurant_manager.migrations import default_state


def test_template_preview_and_apply(tmp_path: Path):
    source = create_import_template(tmp_path / "import.xlsx")
    preview = preview_import(source, default_state())
    assert preview["errors"] == []
    assert preview["counts"] == {"quickExpenses": 1, "duplicateQuickExpenses": 0, "purchaseOrders": 1, "purchaseLines": 2}
    state = apply_import(preview, default_state(), True)
    assert [row["amount"] for row in state["expenses"]] == [680.0, 336.0, 64.0]
    assert len(state["products"]) == 2
    assert len(state["importBatches"]) == 1


def test_money_is_rounded_as_yuan_and_duplicate_order_is_rejected(tmp_path: Path):
    source = create_import_template(tmp_path / "import.xlsx")
    book = load_workbook(source)
    book["快速支出"]["C2"] = 1.235
    book.save(source)
    state = default_state()
    preview = preview_import(source, state)
    assert preview["quickExpenses"][0]["amount"] == 1.24
    state["expenses"].append({"id": 1, "purchaseNo": "CG-202608-001"})
    duplicate = preview_import(source, state)
    assert any("不能重复导入" in error for error in duplicate["errors"])


def test_duplicate_quick_expenses_are_listed_and_skipped_by_default(tmp_path: Path):
    source = create_import_template(tmp_path / "import.xlsx")
    book = load_workbook(source)
    quick = book["快速支出"]
    quick.append([cell.value for cell in quick[2]])
    book.save(source)

    preview = preview_import(source, default_state())

    assert preview["errors"] == []
    assert preview["counts"]["quickExpenses"] == 1
    assert preview["counts"]["duplicateQuickExpenses"] == 1
    assert preview["duplicateQuickExpenses"][0]["reason"] == "文件内重复"
    state = apply_import(preview, default_state(), True)
    assert len([row for row in state["expenses"] if row["mode"] == "快速记账"]) == 1
    assert state["importBatches"][0]["quickExpenses"] == 1


def test_duplicate_quick_expenses_can_be_imported_after_confirmation(tmp_path: Path):
    source = create_import_template(tmp_path / "import.xlsx")
    book = load_workbook(source)
    quick = book["快速支出"]
    quick.append([cell.value for cell in quick[2]])
    book.save(source)

    preview = preview_import(source, default_state())
    state = apply_import(preview, default_state(), True, True)

    assert len([row for row in state["expenses"] if row["mode"] == "快速记账"]) == 2
    assert state["importBatches"][0]["quickExpenses"] == 2
    assert state["importBatches"][0]["duplicateQuickExpenses"] == 1


def test_existing_quick_expense_is_detected_with_normalized_note(tmp_path: Path):
    source = create_import_template(tmp_path / "import.xlsx")
    book = load_workbook(source)
    book["详细采购"].delete_rows(2, 2)
    book.save(source)
    first_preview = preview_import(source, default_state())
    state = apply_import(first_preview, default_state(), True)

    duplicate = preview_import(source, state)

    assert duplicate["errors"] == []
    assert duplicate["counts"]["quickExpenses"] == 0
    assert duplicate["counts"]["duplicateQuickExpenses"] == 1
    assert duplicate["duplicateQuickExpenses"][0]["reason"] == "系统已有相同记录"


def test_legacy_custom_quick_expense_item_matches_import_category(tmp_path: Path):
    source = create_import_template(tmp_path / "legacy-custom.xlsx")
    book = load_workbook(source)
    quick = book["快速支出"]
    quick.delete_rows(2, quick.max_row)
    quick.append(["2026-08-01", "米油", 680, "李经理", ""])
    quick.append(["2026-08-02", "耗材", 128, "王丽", ""])
    book["详细采购"].delete_rows(2, book["详细采购"].max_row)
    book.save(source)

    state = default_state()
    state["expenseCategories"].append({"id": 99, "name": "米油", "active": True})
    state["expenses"].extend([
        {"id": 1, "date": "2026-08-01", "mode": "快速记账", "category": "其他", "item": "米油", "amount": 680, "handler": "李经理", "status": "有效"},
        {"id": 2, "date": "2026-08-02", "mode": "快速支出", "category": "其他", "item": "耗材", "amount": 128, "handler": "王丽", "status": "有效"},
    ])

    preview = preview_import(source, state)

    assert preview["errors"] == []
    assert preview["counts"]["quickExpenses"] == 0
    assert preview["counts"]["duplicateQuickExpenses"] == 2
    assert {row["category"] for row in preview["duplicateQuickExpenses"]} == {"米油", "耗材"}
    assert {row["reason"] for row in preview["duplicateQuickExpenses"]} == {"系统已有相同记录"}


def test_legacy_custom_match_still_requires_same_business_label(tmp_path: Path):
    source = create_import_template(tmp_path / "different-custom.xlsx")
    book = load_workbook(source)
    quick = book["快速支出"]
    quick.delete_rows(2, quick.max_row)
    quick.append(["2026-08-01", "米油", 680, "李经理", ""])
    book["详细采购"].delete_rows(2, book["详细采购"].max_row)
    book.save(source)

    state = default_state()
    state["expenseCategories"].append({"id": 99, "name": "米油", "active": True})
    state["expenses"].append({"id": 1, "date": "2026-08-01", "mode": "快速记账", "category": "其他", "item": "耗材", "amount": 680, "handler": "李经理", "status": "有效"})

    preview = preview_import(source, state)

    assert preview["errors"] == []
    assert preview["counts"]["quickExpenses"] == 1
    assert preview["counts"]["duplicateQuickExpenses"] == 0


def test_imported_quick_expense_preserves_note_field(tmp_path: Path):
    source = create_import_template(tmp_path / "note.xlsx")
    book = load_workbook(source)
    book["详细采购"].delete_rows(2, book["详细采购"].max_row)
    book.save(source)

    preview = preview_import(source, default_state())
    state = apply_import(preview, default_state(), True)

    assert state["expenses"][0]["note"] == "本月水电费"
